from openpyxl import load_workbook

from nomenclator import Nomenclator


class FactoriaNomenclator:
    @classmethod
    def desde_excel(cls, ruta_excel):
        libro = load_workbook(ruta_excel, data_only=True)
        nomenclator = Nomenclator()

        for hoja in libro.worksheets:
            genero = cls._obtener_genero_desde_hoja(hoja.title)
            decadas = cls._leer_decadas(hoja)

            for decada in decadas.values():
                nomenclator.registrar_decada(decada)

            for columna_inicio, decada in decadas.items():
                cls._cargar_bloque_de_decada(hoja, columna_inicio, decada, genero, nomenclator)

        return nomenclator

    @classmethod
    def _obtener_genero_desde_hoja(cls, nombre_hoja):
        nombre = nombre_hoja.strip().upper()
        if nombre == "HOMBRES":
            return "H"
        return "M"

    @classmethod
    def _leer_decadas(cls, hoja):
        decadas = {}
        columna = 2
        while columna <= hoja.max_column:
            valor = hoja.cell(1, columna).value
            if valor is not None:
                decadas[columna] = cls._normalizar_decada(str(valor).strip())
            columna += 3
        return decadas

    @classmethod
    def _cargar_bloque_de_decada(cls, hoja, columna_inicio, decada, genero, nomenclator):
        fila = 3
        while fila <= hoja.max_row:
            nombre_texto = hoja.cell(fila, columna_inicio).value
            frecuencia = hoja.cell(fila, columna_inicio + 1).value
            tanto_por_mil = hoja.cell(fila, columna_inicio + 2).value

            if nombre_texto is not None and frecuencia is not None and tanto_por_mil is not None:
                nombre_limpio = str(nombre_texto).strip()
                if nombre_limpio != "":
                    nombre = nomenclator.obtener_o_crear_nombre(nombre_limpio, genero)
                    nombre.agregar_registro(
                        decada,
                        int(frecuencia),
                        float(tanto_por_mil)
                    )

            fila += 1

    @classmethod
    def _normalizar_decada(cls, texto):
        texto = texto.upper()

        if "ANTES DE 1930" in texto:
            return "ANTES DE 1930"

        if "2020, 2021 Y 2022" in texto:
            return "2020-2022"

        if "1930 A 1939" in texto:
            return "1930-1939"
        if "1940 A 1949" in texto:
            return "1940-1949"
        if "1950 A 1959" in texto:
            return "1950-1959"
        if "1960 A 1969" in texto:
            return "1960-1969"
        if "1970 A 1979" in texto:
            return "1970-1979"
        if "1980 A 1989" in texto:
            return "1980-1989"
        if "1990 A 1999" in texto:
            return "1990-1999"
        if "2000 A 2009" in texto:
            return "2000-2009"
        if "2010 A 2019" in texto:
            return "2010-2019"

        return texto
