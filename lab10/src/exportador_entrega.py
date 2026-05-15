from __future__ import annotations

import csv
import re
import zipfile
from datetime import datetime
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ImagenExcel
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as ImagenPIL


class ExportadorEntrega:
    def exportar(self, data_dir: str | Path) -> tuple[Path, Path]:
        data_dir = Path(data_dir)
        comparacion = self._leer_comparacion(data_dir / "tablas_entrega_lab10.md")
        tablas_mae = self._leer_tablas_mae(data_dir / "resultados_lab10.txt")
        predicciones = self._leer_predicciones(
            data_dir / "predicciones_diferencias_lab10.csv",
        )
        pares, anomalias = self._leer_anomalias(data_dir / "anomalias_lab10.txt")
        graficas = sorted(data_dir.glob("series_*.png"))

        ruta_word = data_dir / "entrega_lab10.docx"
        ruta_excel = data_dir / "entrega_lab10.xlsx"

        self._crear_excel(
            ruta_excel,
            comparacion,
            tablas_mae,
            predicciones,
            pares,
            anomalias,
            graficas,
        )
        self._crear_word(
            ruta_word,
            comparacion,
            tablas_mae,
            pares,
            anomalias,
            graficas,
        )

        return ruta_word, ruta_excel

    def _leer_comparacion(self, ruta: Path) -> list[list[str]]:
        filas: list[list[str]] = []

        for linea in ruta.read_text(encoding="utf-8").splitlines():
            linea = linea.strip()
            if not linea.startswith("|") or "---" in linea:
                continue
            celdas = [celda.strip() for celda in linea.strip("|").split("|")]
            filas.append(celdas)

        return filas

    def _leer_tablas_mae(self, ruta: Path) -> dict[str, list[list[str]]]:
        lineas = ruta.read_text(encoding="utf-8").splitlines()
        tablas: dict[str, list[list[str]]] = {
            "Recta de regresion": [
                ["PH", "Tasa", "Normalizacion", "MAE val", "MAE test", "MAE lab9"],
            ],
            "kNN correlacion": [
                ["PH", "K", "MAE val", "MAE test", "MAE lab9"],
            ],
            "kNN heuristica": [
                ["PH", "K", "MAE val", "MAE test", "MAE lab9"],
            ],
        }
        seccion = ""

        for linea in lineas:
            texto = linea.strip()
            if texto == "RECTA DE REGRESION SOBRE DIFERENCIAS":
                seccion = "Recta de regresion"
                continue
            if texto == "KNN PONDERADO CON CORRELACION SOBRE DIFERENCIAS":
                seccion = "kNN correlacion"
                continue
            if texto == "KNN CON HEURISTICA SOBRE DIFERENCIAS":
                seccion = "kNN heuristica"
                continue
            if texto.startswith("COMPARACION"):
                seccion = ""
                continue
            if not seccion or texto.startswith("PH"):
                continue

            partes = texto.split()
            if seccion == "Recta de regresion" and len(partes) == 6:
                tablas[seccion].append(partes)
            elif seccion in ("kNN correlacion", "kNN heuristica") and len(partes) == 5:
                tablas[seccion].append(partes)

        return tablas

    def _leer_predicciones(self, ruta: Path) -> list[list[str]]:
        with ruta.open("r", encoding="utf-8", newline="") as fichero:
            lector = csv.reader(fichero)
            return [fila for fila in lector]

    def _leer_anomalias(self, ruta: Path) -> tuple[list[tuple[str, str]], list[list[str]]]:
        pares: list[tuple[str, str]] = []
        anomalias: list[list[str]] = [
            [
                "Umbral ruptura",
                "Separacion maxima",
                "Duracion minima",
                "Dia",
                "Variable 1",
                "Variable 2",
                "Inicio",
                "Fin",
                "Duracion",
            ],
        ]
        umbral = ""
        separacion = ""
        duracion_minima = ""
        dia = ""
        leyendo_pares = False

        patron_configuracion = re.compile(
            r"Umbral ruptura=([^,]+), separacion=([^,]+), duracion_minima=(.+)",
        )
        patron_dia = re.compile(r"Dia\s+(\d+):")
        patron_intervalos = re.compile(r"([^/]+)\s+/\s+([^:]+):\s+(.+)")

        for linea in ruta.read_text(encoding="utf-8").splitlines():
            texto = linea.strip()
            if texto.startswith("Pares correlacionados"):
                leyendo_pares = True
                continue
            if texto.startswith("Graficas exportadas"):
                leyendo_pares = False
                continue
            if leyendo_pares and texto.startswith("- "):
                nombre1, nombre2 = texto[2:].split(" / ")
                pares.append((nombre1, nombre2))
                continue

            coincidencia = patron_configuracion.match(texto)
            if coincidencia:
                umbral, separacion, duracion_minima = coincidencia.groups()
                continue

            coincidencia = patron_dia.search(texto)
            if coincidencia:
                dia = coincidencia.group(1)
                continue

            coincidencia = patron_intervalos.match(texto)
            if coincidencia and umbral and dia:
                nombre1, nombre2, texto_intervalos = coincidencia.groups()
                for intervalo in texto_intervalos.split(","):
                    inicio, fin = intervalo.strip().split("-")
                    duracion = int(fin) - int(inicio) + 1
                    anomalias.append(
                        [
                            umbral,
                            separacion,
                            duracion_minima,
                            dia,
                            nombre1.strip(),
                            nombre2.strip(),
                            inicio,
                            fin,
                            str(duracion),
                        ],
                    )

        return pares, anomalias

    def _crear_excel(
        self,
        ruta: Path,
        comparacion: list[list[str]],
        tablas_mae: dict[str, list[list[str]]],
        predicciones: list[list[str]],
        pares: list[tuple[str, str]],
        anomalias: list[list[str]],
        graficas: list[Path],
    ) -> None:
        libro = Workbook()
        hoja_resumen = libro.active
        hoja_resumen.title = "Resumen MAE"

        fila_actual = 1
        fila_actual = self._escribir_bloque_tabla(
            hoja_resumen,
            "Comparacion de mejores modelos",
            comparacion,
            fila_actual,
        )
        for titulo, filas in tablas_mae.items():
            fila_actual = self._escribir_bloque_tabla(
                hoja_resumen,
                titulo,
                filas,
                fila_actual + 1,
            )

        hoja_predicciones = libro.create_sheet("Predicciones")
        self._escribir_filas(hoja_predicciones, predicciones)

        hoja_pares = libro.create_sheet("Pares dia 1")
        self._escribir_filas(
            hoja_pares,
            [["Variable 1", "Variable 2"]]
            + [[nombre1, nombre2] for nombre1, nombre2 in pares],
        )

        hoja_anomalias = libro.create_sheet("Anomalias")
        self._escribir_filas(hoja_anomalias, anomalias)

        hoja_graficas = libro.create_sheet("Graficas")
        hoja_graficas["A1"] = "Graficas exportadas"
        hoja_graficas["A1"].font = Font(bold=True, size=14)
        fila = 3
        for grafica in graficas:
            hoja_graficas[f"A{fila}"] = grafica.name
            imagen = ImagenExcel(str(grafica))
            imagen.width = 620
            imagen.height = 290
            hoja_graficas.add_image(imagen, f"B{fila}")
            fila += 17

        for hoja in libro.worksheets:
            self._aplicar_estilo_hoja(hoja)

        libro.save(ruta)

    def _escribir_bloque_tabla(
        self,
        hoja,
        titulo: str,
        filas: list[list[str]],
        fila_inicio: int,
    ) -> int:
        hoja.cell(row=fila_inicio, column=1, value=titulo)
        hoja.cell(row=fila_inicio, column=1).font = Font(bold=True, size=13)
        fila_actual = fila_inicio + 1
        self._escribir_filas(hoja, filas, fila_actual)
        return fila_actual + len(filas) + 1

    def _escribir_filas(self, hoja, filas: list[list[str]], fila_inicio: int = 1) -> None:
        for indice_fila, fila in enumerate(filas, start=fila_inicio):
            for indice_columna, valor in enumerate(fila, start=1):
                celda = hoja.cell(row=indice_fila, column=indice_columna, value=valor)
                if indice_fila == fila_inicio:
                    celda.font = Font(bold=True)
                    celda.fill = PatternFill("solid", fgColor="D9EAF7")

    def _aplicar_estilo_hoja(self, hoja) -> None:
        hoja.freeze_panes = "A2"
        for columna in range(1, hoja.max_column + 1):
            letra = get_column_letter(columna)
            ancho = 14
            for celda in hoja[letra]:
                if celda.value is not None:
                    ancho = max(ancho, min(len(str(celda.value)) + 2, 45))
                celda.alignment = Alignment(vertical="top", wrap_text=True)
            hoja.column_dimensions[letra].width = ancho

    def _crear_word(
        self,
        ruta: Path,
        comparacion: list[list[str]],
        tablas_mae: dict[str, list[list[str]]],
        pares: list[tuple[str, str]],
        anomalias: list[list[str]],
        graficas: list[Path],
    ) -> None:
        relaciones: list[str] = []
        partes: list[str] = [
            self._parrafo(
                "Laboratorio 10 - Entrega de resultados",
                tamano=32,
                negrita=True,
            ),
            self._parrafo(
                "Generado el "
                + datetime.now().strftime("%d/%m/%Y %H:%M")
                + ". Los MAE de diferencias estan reconstruidos sobre la serie original.",
            ),
            self._parrafo("Resumen MAE", tamano=26, negrita=True),
            self._tabla_word(comparacion),
            self._parrafo("MAE completos", tamano=26, negrita=True),
        ]

        for titulo, filas in tablas_mae.items():
            partes.append(self._parrafo(titulo, tamano=23, negrita=True))
            partes.append(self._tabla_word(filas))

        partes.extend(
            [
                self._parrafo("Deteccion de anomalias", tamano=26, negrita=True),
                self._parrafo(
                    f"Pares correlacionados en el dia 1 con umbral 0.8: {len(pares)}.",
                ),
                self._tabla_word(anomalias),
                self._parrafo("Representaciones graficas", tamano=26, negrita=True),
            ],
        )

        for indice, grafica in enumerate(graficas, start=1):
            rid = f"rId{indice}"
            nombre_media = f"image{indice}.png"
            relaciones.append(
                '<Relationship Id="'
                + rid
                + '" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="media/'
                + nombre_media
                + '"/>',
            )
            partes.append(self._parrafo(grafica.name, negrita=True))
            partes.append(self._imagen_word(rid, indice, grafica))

        document_xml = self._documento_word("".join(partes))
        relaciones_xml = self._relaciones_word("".join(relaciones))

        with zipfile.ZipFile(ruta, "w", compression=zipfile.ZIP_DEFLATED) as paquete:
            paquete.writestr("[Content_Types].xml", self._tipos_contenido())
            paquete.writestr("_rels/.rels", self._relaciones_raiz())
            paquete.writestr("docProps/core.xml", self._core_xml())
            paquete.writestr("docProps/app.xml", self._app_xml())
            paquete.writestr("word/document.xml", document_xml)
            paquete.writestr("word/_rels/document.xml.rels", relaciones_xml)
            for indice, grafica in enumerate(graficas, start=1):
                paquete.write(grafica, f"word/media/image{indice}.png")

    def _parrafo(
        self,
        texto: str,
        tamano: int = 21,
        negrita: bool = False,
    ) -> str:
        negrita_xml = "<w:b/>" if negrita else ""
        return (
            "<w:p><w:pPr><w:spacing w:after=\"120\"/></w:pPr><w:r><w:rPr>"
            + negrita_xml
            + f"<w:sz w:val=\"{tamano}\"/></w:rPr><w:t>"
            + escape(texto)
            + "</w:t></w:r></w:p>"
        )

    def _tabla_word(self, filas: list[list[str]]) -> str:
        tabla = [
            "<w:tbl><w:tblPr><w:tblBorders>"
            "<w:top w:val=\"single\" w:sz=\"4\" w:space=\"0\" w:color=\"A6A6A6\"/>"
            "<w:left w:val=\"single\" w:sz=\"4\" w:space=\"0\" w:color=\"A6A6A6\"/>"
            "<w:bottom w:val=\"single\" w:sz=\"4\" w:space=\"0\" w:color=\"A6A6A6\"/>"
            "<w:right w:val=\"single\" w:sz=\"4\" w:space=\"0\" w:color=\"A6A6A6\"/>"
            "<w:insideH w:val=\"single\" w:sz=\"4\" w:space=\"0\" w:color=\"A6A6A6\"/>"
            "<w:insideV w:val=\"single\" w:sz=\"4\" w:space=\"0\" w:color=\"A6A6A6\"/>"
            "</w:tblBorders></w:tblPr>"
        ]
        for indice_fila, fila in enumerate(filas):
            tabla.append("<w:tr>")
            for celda in fila:
                negrita = "<w:b/>" if indice_fila == 0 else ""
                tabla.append(
                    "<w:tc><w:tcPr><w:tcW w:w=\"0\" w:type=\"auto\"/></w:tcPr>"
                    "<w:p><w:r><w:rPr>"
                    + negrita
                    + "<w:sz w:val=\"18\"/></w:rPr><w:t>"
                    + escape(str(celda))
                    + "</w:t></w:r></w:p></w:tc>",
                )
            tabla.append("</w:tr>")
        tabla.append("</w:tbl>")
        tabla.append(self._parrafo(""))
        return "".join(tabla)

    def _imagen_word(self, rid: str, indice: int, grafica: Path) -> str:
        ancho_emu = int(6.4 * 914400)
        with ImagenPIL.open(grafica) as imagen:
            ratio = imagen.height / imagen.width
        alto_emu = int(ancho_emu * ratio)

        return f"""
<w:p>
  <w:r>
    <w:drawing>
      <wp:inline distT="0" distB="0" distL="0" distR="0">
        <wp:extent cx="{ancho_emu}" cy="{alto_emu}"/>
        <wp:docPr id="{indice}" name="{escape(grafica.name)}"/>
        <wp:cNvGraphicFramePr>
          <a:graphicFrameLocks noChangeAspect="1"/>
        </wp:cNvGraphicFramePr>
        <a:graphic>
          <a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/picture">
            <pic:pic>
              <pic:nvPicPr>
                <pic:cNvPr id="{indice}" name="{escape(grafica.name)}"/>
                <pic:cNvPicPr/>
              </pic:nvPicPr>
              <pic:blipFill>
                <a:blip r:embed="{rid}"/>
                <a:stretch><a:fillRect/></a:stretch>
              </pic:blipFill>
              <pic:spPr>
                <a:xfrm>
                  <a:off x="0" y="0"/>
                  <a:ext cx="{ancho_emu}" cy="{alto_emu}"/>
                </a:xfrm>
                <a:prstGeom prst="rect"><a:avLst/></a:prstGeom>
              </pic:spPr>
            </pic:pic>
          </a:graphicData>
        </a:graphic>
      </wp:inline>
    </w:drawing>
  </w:r>
</w:p>
"""

    def _documento_word(self, cuerpo: str) -> str:
        return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:wpc="http://schemas.microsoft.com/office/word/2010/wordprocessingCanvas"
 xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006"
 xmlns:o="urn:schemas-microsoft-com:office:office"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
 xmlns:m="http://schemas.openxmlformats.org/officeDocument/2006/math"
 xmlns:v="urn:schemas-microsoft-com:vml"
 xmlns:wp14="http://schemas.microsoft.com/office/word/2010/wordprocessingDrawing"
 xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing"
 xmlns:w10="urn:schemas-microsoft-com:office:word"
 xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"
 xmlns:w14="http://schemas.microsoft.com/office/word/2010/wordml"
 xmlns:wpg="http://schemas.microsoft.com/office/word/2010/wordprocessingGroup"
 xmlns:wpi="http://schemas.microsoft.com/office/word/2010/wordprocessingInk"
 xmlns:wne="http://schemas.microsoft.com/office/word/2006/wordml"
 xmlns:wps="http://schemas.microsoft.com/office/word/2010/wordprocessingShape"
 xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
 xmlns:pic="http://schemas.openxmlformats.org/drawingml/2006/picture"
 mc:Ignorable="w14 wp14">
  <w:body>
    {cuerpo}
    <w:sectPr>
      <w:pgSz w:w="11906" w:h="16838"/>
      <w:pgMar w:top="720" w:right="720" w:bottom="720" w:left="720"/>
    </w:sectPr>
  </w:body>
</w:document>"""

    def _relaciones_word(self, relaciones: str) -> str:
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            + relaciones
            + "</Relationships>"
        )

    def _relaciones_raiz(self) -> str:
        return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>
</Relationships>"""

    def _tipos_contenido(self) -> str:
        return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Default Extension="png" ContentType="image/png"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
  <Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
  <Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>
</Types>"""

    def _core_xml(self) -> str:
        fecha = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
        return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties"
 xmlns:dc="http://purl.org/dc/elements/1.1/"
 xmlns:dcterms="http://purl.org/dc/terms/"
 xmlns:dcmitype="http://purl.org/dc/dcmitype/"
 xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:title>Laboratorio 10 - Entrega de resultados</dc:title>
  <dc:creator>Codex</dc:creator>
  <cp:lastModifiedBy>Codex</cp:lastModifiedBy>
  <dcterms:created xsi:type="dcterms:W3CDTF">{fecha}</dcterms:created>
  <dcterms:modified xsi:type="dcterms:W3CDTF">{fecha}</dcterms:modified>
</cp:coreProperties>"""

    def _app_xml(self) -> str:
        return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"
 xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
  <Application>Microsoft Office Word</Application>
</Properties>"""
