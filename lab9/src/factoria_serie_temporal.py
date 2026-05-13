from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from dataset_regresion import DataSetRegresion
from observacion_serie_temporal import ObservacionSerieTemporal
from registro_regresion import RegistroRegresion
from serie_temporal import SerieTemporal


class FactoriaSerieTemporal:
    @classmethod
    def leer_excel(cls, ruta_excel: str | Path) -> SerieTemporal:
        ruta = Path(ruta_excel)
        if not ruta.exists():
            raise FileNotFoundError(f"No existe el fichero: {ruta}")

        libro = load_workbook(ruta, data_only=True, read_only=True)
        hoja = libro[libro.sheetnames[0]]
        serie = SerieTemporal()

        for fila in hoja.iter_rows(min_row=2, values_only=True):
            fecha = fila[0] if len(fila) > 0 else None
            valor = fila[1] if len(fila) > 1 else None

            if fecha is None or valor is None:
                continue

            serie.agregar_observacion(ObservacionSerieTemporal(fecha, float(valor)))

        return serie

    @classmethod
    def crear_dataset_regresion_desde_excel(
        cls,
        ruta_excel: str | Path,
        ph: int,
        max_registros: int | None = None,
    ) -> DataSetRegresion:
        serie = cls.leer_excel(ruta_excel)
        return cls.crear_dataset_regresion_desde_serie(serie, ph, max_registros=max_registros)

    @classmethod
    def crear_dataset_regresion_desde_serie(
        cls,
        serie: SerieTemporal,
        ph: int,
        max_registros: int | None = None,
    ) -> DataSetRegresion:
        if ph <= 0:
            raise ValueError("El valor de PH debe ser mayor que cero.")

        valores = serie.valores()
        if len(valores) <= ph:
            raise ValueError("La serie no tiene suficientes valores para crear el dataset.")

        dataset = DataSetRegresion()
        cabeceras = [f"Lag-{indice}" for indice in range(ph, 0, -1)]
        dataset.establecer_cabeceras(cabeceras, "objetivo")

        for indice in range(ph, len(valores)):
            atributos = valores[indice - ph : indice]
            objetivo = valores[indice]
            dataset.agregar_registro(RegistroRegresion(atributos, objetivo))

        return dataset.ultimos_registros(max_registros)

