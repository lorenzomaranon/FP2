from openpyxl import load_workbook

from estacion import Estacion
from linea import Linea
from lineas import Lineas


class FactoriaMetro:
    def leer_lineas_desde_excel(self, ruta_excel):
        libro = load_workbook(ruta_excel, data_only=True)
        hoja = libro.active
        lineas = Lineas()
        estaciones_por_nombre = {}
        errores = []

        fila = 2
        while fila <= hoja.max_row:
            nombre_linea = hoja.cell(fila, 1).value
            texto_estaciones = hoja.cell(fila, 2).value
            total_paradas = hoja.cell(fila, 3).value

            if nombre_linea is not None and texto_estaciones is not None:
                nombres_estaciones = self._separar_estaciones(texto_estaciones)
                es_circular = self._es_lista_circular(nombres_estaciones)

                if es_circular:
                    nombres_estaciones = nombres_estaciones[0:len(nombres_estaciones) - 1]

                if total_paradas is not None and len(nombres_estaciones) != int(total_paradas):
                    mensaje = str(nombre_linea) + ": el Excel indica "
                    mensaje = mensaje + str(total_paradas) + " paradas, pero se han leido "
                    mensaje = mensaje + str(len(nombres_estaciones)) + "."
                    errores.append(mensaje)

                linea = Linea(nombre_linea, es_circular)
                estaciones = []

                for nombre_estacion in nombres_estaciones:
                    clave = self._normalizar_nombre(nombre_estacion)

                    if clave not in estaciones_por_nombre:
                        estaciones_por_nombre[clave] = Estacion(nombre_estacion)

                    estaciones.append(estaciones_por_nombre[clave])

                lineas.agregar_linea(linea, estaciones)

            fila = fila + 1

        if len(errores) > 0:
            mensaje_error = "Se han encontrado errores en el Excel:\n"
            for error in errores:
                mensaje_error = mensaje_error + "- " + error + "\n"
            raise ValueError(mensaje_error)

        return lineas

    def _separar_estaciones(self, texto_estaciones):
        partes = str(texto_estaciones).split(",")
        estaciones = []

        for parte in partes:
            nombre = parte.strip()
            if nombre != "":
                estaciones.append(nombre)

        return estaciones

    def _es_lista_circular(self, nombres_estaciones):
        if len(nombres_estaciones) < 2:
            return False

        primera = self._normalizar_nombre(nombres_estaciones[0])
        ultima = self._normalizar_nombre(nombres_estaciones[len(nombres_estaciones) - 1])

        return primera == ultima

    def _normalizar_nombre(self, nombre):
        return str(nombre).strip().upper()
